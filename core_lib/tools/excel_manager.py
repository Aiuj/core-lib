"""
Module for loading and manipulating Excel files, and rendering their contents as markdown tables.
"""

from openpyxl import load_workbook
from tabulate import tabulate
from core_lib import get_module_logger
import os

# Get module-specific logger
logger = get_module_logger()

class ExcelManager:
    """
    Manages loading, rendering, and manipulation of Excel workbooks.
    """

    def __init__(self, excel_path: str = None, excel_bytes: bytes = None):
        """
        Initialize the ExcelManager with either a path to the Excel file or raw bytes.

        Args:
            excel_path (str, optional): Path to the Excel file.
            excel_bytes (bytes, optional): Raw bytes of an Excel file. If provided, workbook
                will be loaded from memory and no temporary file is required.
        """
        self.excel_path = excel_path
        self.excel_bytes = excel_bytes
        self._bytes_io = None
        self.wb = None
        logger.debug(f"ExcelManager initialized with path: {excel_path} and bytes: {'present' if excel_bytes else 'none'}")

    def load(self):
        """
        Loads the Excel workbook from the specified path.

        Returns:
            Workbook: The loaded openpyxl Workbook object.
        """
        try:
            if self.excel_bytes is not None:
                # Load from in-memory bytes without creating a temporary file
                from io import BytesIO
                logger.info("Loading Excel workbook from in-memory bytes")
                self._bytes_io = BytesIO(self.excel_bytes)
                self.wb = load_workbook(self._bytes_io, read_only=True)
            else:
                logger.info(f"Loading Excel workbook from: {self.excel_path}")
                self.wb = load_workbook(self.excel_path, read_only=True)
            logger.info(f"Excel workbook loaded successfully - sheets: {self.wb.sheetnames}")
            return self.wb
        except Exception as e:
            logger.error(f"Failed to load Excel workbook from {self.excel_path}: {str(e)}")
            raise

    def close(self):
        """
        Close any open workbook and release associated resources (like BytesIO).
        Safe to call multiple times.
        """
        try:
            if self.wb is not None:
                try:
                    # openpyxl Workbook objects expose a close method on read-only mode
                    close_fn = getattr(self.wb, 'close', None)
                    if callable(close_fn):
                        close_fn()
                except Exception:
                    # ignore errors closing workbook
                    pass
                self.wb = None

            if self._bytes_io is not None:
                try:
                    self._bytes_io.close()
                except Exception:
                    pass
                self._bytes_io = None
        except Exception:
            # Ensure we don't raise from cleanup
            pass

    def __del__(self):
        try:
            self.close()
        except Exception:
            pass

    # Replace None and NaN with empty string
    def clean_cell(self, cell):
        if cell is None:
            return ''
        try:
            if isinstance(cell, float) and str(cell) == 'nan':
                return ''
        except Exception:
            pass
        return cell

    def get_sheet_tables(self, ws, max_rows=None, add_col_headers=False, add_row_headers=False):
        """
        Extracts sheet data as a list of lists, with options to add Excel-style column and row headers.
        Preserves original cell locations for headers, even if empty rows/columns are removed.

        Args:
            ws: Worksheet object.
            max_rows (int, optional): Maximum number of rows to return.
            add_col_headers (bool): If True, adds Excel-style column headers (A, B, ...).
            add_row_headers (bool): If True, adds Excel-style row headers (1, 2, ...).

        Returns:
            list: Sheet data as a list of lists, with optional headers.
        """
        # Get all rows as lists, preserving original shape
        data = list(ws.values)
        num_rows = len(data)
        num_cols = max((len(row) for row in data), default=0)
        # Normalize all rows to same length
        norm_data = [list(row) + [''] * (num_cols - len(row)) for row in data]

        # Track which rows/cols are empty
        empty_row_flags = [not any(cell not in (None, '', float('nan')) for cell in row) for row in norm_data]
        empty_col_flags = [not any(norm_data[row_idx][col_idx] not in (None, '', float('nan')) for row_idx in range(num_rows)) for col_idx in range(num_cols)]

        # Remove empty rows/cols but keep original indices for headers
        row_map = [i for i, empty in enumerate(empty_row_flags) if not empty]
        col_map = [i for i, empty in enumerate(empty_col_flags) if not empty]
        filtered_data = [[norm_data[row_idx][col_idx] for col_idx in col_map] for row_idx in row_map]

        # Limit rows
        if max_rows is not None:
            filtered_data = filtered_data[:max_rows]
            row_map = row_map[:max_rows]

        # Clean cells
        filtered_data = [[self.clean_cell(cell) for cell in row] for row in filtered_data]

        # Add column headers (A, B, C, ...)
        if add_col_headers and filtered_data:
            def excel_col_name(n):
                name = ''
                while n > 0:
                    n, r = divmod(n-1, 26)
                    name = chr(65 + r) + name
                return name
            col_headers = [excel_col_name(i+1) for i in col_map]
            if add_row_headers:
                col_headers = [''] + col_headers
            filtered_data.insert(0, col_headers)

        # Add row headers (1, 2, ...)
        if add_row_headers and filtered_data:
            for idx, row in enumerate(filtered_data):
                # If col headers are present, skip the first row
                if add_col_headers and idx == 0:
                    continue
                # Use original Excel row number (1-based)
                row_num = str(row_map[idx-1]+1) if add_col_headers else str(row_map[idx]+1)
                row.insert(0, row_num)

        return filtered_data
    
    def get_content(self, max_rows=None, add_col_headers=True, add_row_headers=True, detect_language=True):
        """
        Returns the content of the workbook as a markdown string.

        Args:
            max_rows (int, optional): Maximum number of rows to render per sheet.
            add_col_headers (bool): If True, adds Excel-style column headers (A, B, ...).
            add_row_headers (bool): If True, adds Excel-style row headers (1, 2, ...).
            detect_language (bool): If True, detects language of the content.

        Returns:
            list: List of dictionaries containing sheet name, markdown content, language, and rows.
        
        Raises:
            ValueError: If the workbook is not loaded.
        """
        from core_lib.utils.language_utils import LanguageUtils

        if self.wb is None:
            logger.error("Workbook not loaded. Cannot get content.")
            raise ValueError("Workbook not loaded. Call load() first.")

        logger.info(f"Getting content from workbook - max_rows: {max_rows}")
        results = []
        
        for sheet_name in self.wb.sheetnames:
            logger.debug(f"Processing sheet: {sheet_name}")
            ws = self.wb[sheet_name]
            data = self.get_sheet_tables(ws, max_rows, add_col_headers=add_col_headers, add_row_headers=add_row_headers)
            # Prepare headers
            headers = data[0] if data else []
            rows = data[1:] if len(data) > 1 else []
            md_table = tabulate(rows, headers=headers, tablefmt='github')
            
            if detect_language:
                # Detect language using core-lib
                language = LanguageUtils.detect_language(sheet_name + ":\n" + md_table)
                logger.debug(f"Language detected for sheet '{sheet_name}': {language}")
            else:
                language = None
                
            results.append({
                "sheet_name": sheet_name,
                "markdown": md_table,
                "language": language,
                "rows": rows,
            })

        logger.info(f"Content processing completed for {len(results)} sheets")
        return results


    def to_markdown(self, max_rows=None, add_col_headers=True, add_row_headers=True, detect_language=True):
        """
        Converts all sheets in the loaded workbook to a list of markdown tables.

        Args:
            max_rows (int, optional): Maximum number of rows to render per sheet. Renders all rows if None.
            add_col_headers (bool): If True, adds Excel-style column headers (A, B, ...).
            add_row_headers (bool): If True, adds Excel-style row headers (1, 2, ...).
            detect_language (bool): If True, detects language of the content.

        Returns:
            list: List of dictionaries with keys:
                - 'sheet_name': Name of the sheet/tab
                - 'markdown': Markdown-formatted table for the sheet
                - 'language': Detected language (if detect_language=True)
        
        Raises:
            ValueError: If the workbook is not loaded.
        """
        from core_lib.utils.language_utils import LanguageUtils

        if self.wb is None:
            logger.error("Workbook not loaded. Cannot convert to markdown.")
            raise ValueError("Workbook not loaded. Call load() first.")
            
        logger.info(f"Converting workbook to markdown - max_rows: {max_rows}")
        results = []
        
        for sheet_name in self.wb.sheetnames:
            logger.debug(f"Converting sheet to markdown: {sheet_name}")
            ws = self.wb[sheet_name]
            data = self.get_sheet_tables(ws, max_rows, add_col_headers=add_col_headers, add_row_headers=add_row_headers)
            # Prepare headers
            headers = data[0] if data else []
            rows = data[1:] if len(data) > 1 else []
            md_table = tabulate(rows, headers=headers, tablefmt='github')
            
            # Detect language if requested
            language = None
            if detect_language:
                try:
                    language = LanguageUtils.detect_language(sheet_name + ":\n" + md_table)
                    logger.debug(f"Language detected for sheet '{sheet_name}': {language}")
                except Exception as e:
                    logger.warning(f"Language detection failed for sheet '{sheet_name}': {e}")
                    language = None
            
            results.append({
                "sheet_name": sheet_name,
                "markdown": md_table,
                "language": language,
                "row_count": len(rows)
            })
            
        logger.info(f"Markdown conversion completed for {len(self.wb.sheetnames)} sheets")
        return results

    def to_combined_markdown(self, max_rows=None, add_col_headers=True, add_row_headers=True, detect_language=True):
        """
        Converts all sheets in the loaded workbook to a single combined markdown string with titles per tab.

        Args:
            max_rows (int, optional): Maximum number of rows to render per sheet. Renders all rows if None.
            add_col_headers (bool): If True, adds Excel-style column headers (A, B, ...).
            add_row_headers (bool): If True, adds Excel-style row headers (1, 2, ...).
            detect_language (bool): If True, detects language of the content.

        Returns:
            str: Combined markdown-formatted string with all sheets, each with its own title.
        
        Raises:
            ValueError: If the workbook is not loaded.
        """
        logger.info("Converting workbook to combined markdown")
        
        # Get individual sheet markdowns
        sheet_markdowns = self.to_markdown(
            max_rows=max_rows,
            add_col_headers=add_col_headers,
            add_row_headers=add_row_headers,
            detect_language=detect_language
        )
        
        # Combine them with titles
        combined_md = ''
        for sheet_data in sheet_markdowns:
            sheet_name = sheet_data['sheet_name']
            markdown = sheet_data['markdown']
            
            combined_md += f'## {sheet_name}\n\n'
            combined_md += markdown + '\n\n'
        
        logger.info(f"Combined markdown created for {len(sheet_markdowns)} sheets")
        return combined_md

    def to_json_ir(self, filename: str = None, max_rows: int = None) -> dict:
        """
        Build a structured JSON Intermediate Representation (IR) for the loaded workbook.

        Args:
            filename (str, optional): Source filename used for id/metadata.
            max_rows (int, optional): Maximum number of data rows (excluding header) to include per sheet.

        Returns:
            dict: { 'document': { ... }, 'language': detected_language }
        """
        from openpyxl.utils import range_boundaries, get_column_letter
        from tabulate import tabulate
        import uuid
        from core_lib.utils.language_utils import LanguageUtils

        if self.wb is None:
            logger.error("Workbook not loaded. Cannot build JSON IR.")
            raise ValueError("Workbook not loaded. Call load() first.")

        base_name = os.path.splitext(os.path.basename(filename or self.excel_path))[0]
        short_id = uuid.uuid4().hex[:8]

        document = {
            "id": f"{base_name}_{short_id}",
            "type": "excel",
            "source_filename": os.path.basename(filename or self.excel_path),
            "language": None,
            "sheets": []
        }

        all_md_snippets = []

        full_wb = None

        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            
            # Get initial bounds from worksheet dimension (fast)
            min_col = min_row = max_col = max_row = None
            used_cols = set()
            
            try:
                dim = ws.calculate_dimension(force=True)
                from openpyxl.utils import range_boundaries
                min_col, min_row, max_col, max_row = range_boundaries(dim)
            except Exception:
                # Fallback: use worksheet properties
                try:
                    min_col = ws.min_column or 1
                    min_row = ws.min_row or 1
                    max_col = ws.max_column or 1
                    max_row = ws.max_row or 1
                except Exception:
                    min_col = min_row = 1
                    max_col = max_row = 100  # Reasonable default
            
            # Validate bounds are reasonable
            if not all(isinstance(v, int) and v >= 1 for v in [min_col, min_row, max_col, max_row]):
                min_col = min_row = 1
                max_col = max_row = 100
            
            # Limit scan to reasonable bounds to avoid hanging on huge sparse sheets
            MAX_SCAN_ROWS = 500  # Scan first N rows to detect used columns
            MAX_SCAN_COLS = 100  # Reasonable max columns to check
            
            actual_max_col = min(max_col, min_col + MAX_SCAN_COLS - 1) if max_col else min_col + MAX_SCAN_COLS - 1
            actual_max_row = min(max_row, min_row + MAX_SCAN_ROWS - 1) if max_row else min_row + MAX_SCAN_ROWS - 1
            
            # Fast column detection: sample rows to find which columns have data
            try:
                for row in ws.iter_rows(min_row=min_row, max_row=actual_max_row, 
                                        min_col=min_col, max_col=actual_max_col, 
                                        values_only=True):
                    for c_idx, val in enumerate(row, start=min_col):
                        val = self.clean_cell(val)
                        if val not in (None, ''):
                            used_cols.add(c_idx)
            except Exception:
                pass
            
            # Also check if there's data beyond our scan range by sampling a few distant cells
            if max_col and max_col > actual_max_col:
                try:
                    # Sample the last few columns
                    for sample_col in range(max(actual_max_col + 1, max_col - 5), max_col + 1):
                        for sample_row in [min_row, min_row + 1, min_row + 5] if max_row > min_row + 5 else range(min_row, min(min_row + 3, max_row + 1)):
                            try:
                                val = self.clean_cell(ws.cell(row=sample_row, column=sample_col).value)
                                if val not in (None, ''):
                                    used_cols.add(sample_col)
                                    break
                            except Exception:
                                pass
                except Exception:
                    pass

            if all(isinstance(v, int) and v >= 1 for v in [min_col, min_row, max_col, max_row]):
                dim = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
            else:
                min_col = min_row = max_col = max_row = 1
                dim = 'A1:A1'

            columns = {}
            headers_for_md = []
            columns_to_include = []  # Track which columns to include in output
            
            for col_idx in range(min_col, max_col + 1):
                col_letter = get_column_letter(col_idx)
                _ws_for_header = None
                try:
                    _ws_for_header = full_wb[sheet_name] if full_wb is not None else ws
                except Exception:
                    _ws_for_header = ws
                raw_header = _ws_for_header.cell(row=min_row, column=col_idx).value
                header_val = self.clean_cell(raw_header)
                
                has_data = col_idx in used_cols
                has_explicit_header = isinstance(header_val, str) and header_val.strip() != ''
                
                # Only include columns that have data OR have an explicit header
                if has_data or has_explicit_header:
                    if not has_explicit_header:
                        header_val = f"Column {col_letter}"
                    columns[col_letter] = {
                        "header": header_val,
                        "has_data": has_data
                    }
                    columns_to_include.append(col_idx)
                    headers_for_md.append(header_val)

            data_rows = []
            md_rows = []
            max_data_rows = (max_rows if max_rows is not None else (max_row - min_row))
            current_count = 0
            
            # Use iter_rows for efficient batch reading instead of cell-by-cell access
            # Limit the row range to avoid iterating unnecessarily
            row_limit = min(max_row, min_row + max_data_rows + 100)  # +100 buffer for empty rows
            
            # Build a set of column indices we're including for fast lookup
            columns_to_include_set = set(columns_to_include)
            
            try:
                for row_idx, row in enumerate(
                    ws.iter_rows(min_row=min_row + 1, max_row=row_limit,
                                 min_col=min_col, max_col=max_col,
                                 values_only=True),
                    start=min_row + 1
                ):
                    if current_count >= max_data_rows:
                        break
                    
                    cells_obj = {}
                    row_vals_md = []
                    is_non_empty = False
                    
                    for col_offset, val in enumerate(row):
                        col_idx = min_col + col_offset
                        
                        # Skip columns we're not including
                        if col_idx not in columns_to_include_set:
                            continue
                            
                        col_letter = get_column_letter(col_idx)
                        val = self.clean_cell(val)
                        if val not in (None, ''):
                            is_non_empty = True
                            cells_obj[col_letter] = val
                        row_vals_md.append(val)
                    
                    if not is_non_empty:
                        continue
                    
                    data_rows.append({
                        "row": row_idx,
                        "cells": cells_obj
                    })
                    md_rows.append(row_vals_md)
                    current_count += 1
            except Exception as e:
                logger.warning(f"Error iterating rows for sheet {sheet_name}: {e}")

            md_table = tabulate(md_rows, headers=headers_for_md, tablefmt='github') if headers_for_md else ''
            all_md_snippets.append(f"## {sheet_name}\n{md_table}")

            sheet_lang = None
            if md_table:
                try:
                    sheet_lang = LanguageUtils.detect_language(sheet_name + "\n" + md_table)
                except Exception:
                    sheet_lang = None

            block = {
                "block_id": f"b{len(document['sheets']) + 1}",
                "type": "table",
                "range": dim,
                "header_row": min_row,
                "columns": columns,
                "rows": data_rows,
                "text_md": md_table,
                "lang": sheet_lang,
            }

            document["sheets"].append({
                "name": sheet_name,
                "blocks": [block]
            })

        combined_md = "\n\n".join(all_md_snippets)
        try:
            overall_lang = LanguageUtils.detect_language(combined_md) if combined_md.strip() else None
        except Exception:
            overall_lang = None
        document["language"] = overall_lang

        result = {
            "document": document,
            "language": overall_lang,
        }

        logger.info("Workbook JSON IR built successfully")
        return result

